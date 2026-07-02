import csv
import re
from collections import Counter
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from app.schemas.datasets import ColumnProfile, DataQualityIssue, DatasetProfile


SAMPLE_LIMIT = 5
TRANSACTION_FIELDS = {"amount", "transactiondate", "date", "description", "account", "direction", "debit", "credit"}
BUDGET_FIELDS = {"budget", "plannedamount", "actualamount", "variance", "department", "period"}
INVOICE_FIELDS = {"invoice", "invoicenumber", "supplier", "vendor", "duedate", "paymentstatus"}
DATE_FIELDS = {"date", "transactiondate", "occurredat", "duedate", "invoicedate", "paymentdate"}
CURRENCY_FIELDS = {"currency", "currencycode"}


def _normalise_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def _infer_type(values: Iterable[Any]) -> str:
    populated = [value for value in values if value not in (None, "")]
    if not populated:
        return "unknown"
    if all(isinstance(value, (datetime, date)) for value in populated):
        return "date"
    if all(isinstance(value, (int, float, Decimal)) and not isinstance(value, bool) for value in populated):
        return "number"
    strings = [str(value).strip() for value in populated]
    if all(re.fullmatch(r"[-+]?\d+(?:\.\d+)?", value.replace(",", "")) for value in strings):
        return "number"
    if all(_is_valid_date(value) for value in strings):
        return "date"
    return "string"


def _is_valid_date(value: str) -> bool:
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            datetime.strptime(value, pattern)
            return True
        except ValueError:
            continue
    return False


def _dataset_kind(headers: list[str]) -> str:
    names = {_normalise_name(header) for header in headers}
    scores = {
        "transactions": len(names & TRANSACTION_FIELDS),
        "budgets": len(names & BUDGET_FIELDS),
        "invoices": len(names & INVOICE_FIELDS),
    }
    kind, score = max(scores.items(), key=lambda item: item[1])
    return kind if score >= 2 else "general"


def _profile_rows(headers: list[str], rows: list[list[Any]], warnings: list[str] | None = None) -> DatasetProfile:
    issues: list[DataQualityIssue] = []
    columns: list[ColumnProfile] = []
    duplicate_count = len(rows) - len({tuple(str(value) for value in row) for row in rows})
    if duplicate_count:
        issues.append(DataQualityIssue(code="duplicate_rows", severity="warning", message=f"Found {duplicate_count} duplicate data row(s).", affected_rows=duplicate_count))

    for index, header in enumerate(headers):
        name = header or f"column_{index + 1}"
        values = [row[index] if index < len(row) else None for row in rows]
        populated = [value for value in values if value not in (None, "")]
        null_count = len(values) - len(populated)
        if null_count:
            issues.append(DataQualityIssue(code="missing_values", severity="warning", message=f"Column '{name}' contains {null_count} missing value(s).", column=name, affected_rows=null_count))
        normalised = _normalise_name(name)
        if normalised in DATE_FIELDS:
            invalid_dates = sum(1 for value in populated if not isinstance(value, (datetime, date)) and not _is_valid_date(str(value).strip()))
            if invalid_dates:
                issues.append(DataQualityIssue(code="invalid_dates", severity="error", message=f"Column '{name}' contains {invalid_dates} unrecognised date value(s).", column=name, affected_rows=invalid_dates))
        if normalised in CURRENCY_FIELDS:
            currencies = {str(value).strip().upper() for value in populated}
            if len(currencies) > 1:
                issues.append(DataQualityIssue(code="mixed_currencies", severity="warning", message=f"Column '{name}' contains multiple currencies: {', '.join(sorted(currencies))}.", column=name, affected_rows=len(populated)))
        columns.append(ColumnProfile(name=name, data_type=_infer_type(values), non_null_count=len(populated), null_count=null_count, unique_count=len({str(value) for value in populated}), sample_values=populated[:SAMPLE_LIMIT]))

    duplicate_headers = len(headers) - len({_normalise_name(header) for header in headers})
    if duplicate_headers:
        issues.append(DataQualityIssue(code="duplicate_columns", severity="error", message="The file contains duplicate column names."))
    return DatasetProfile(dataset_kind=_dataset_kind(headers), row_count=len(rows), column_count=len(headers), columns=columns, warnings=warnings or [], quality_issues=issues)


def profile_csv(path: Path) -> DatasetProfile:
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        sample = source.read(8192)
        source.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(source, dialect)
        try:
            headers = next(reader)
        except StopIteration:
            return DatasetProfile(row_count=0, column_count=0, warnings=["The CSV file is empty."])
        rows = list(reader)
    widths = Counter(len(row) for row in rows)
    warnings = [] if not widths or set(widths) == {len(headers)} else ["Some rows have a different number of fields than the header."]
    return _profile_rows(headers, rows, warnings)


def profile_xlsx(path: Path) -> DatasetProfile:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(value or "") for value in next(iterator, ())]
    rows = [list(row) for row in iterator]
    profile = _profile_rows(headers, rows)
    profile.sheet_names = workbook.sheetnames
    if len(workbook.sheetnames) > 1:
        profile.warnings.append(f"Profiled the first worksheet only: {workbook.sheetnames[0]}.")
    workbook.close()
    return profile


def profile_parquet(path: Path) -> DatasetProfile:
    import pyarrow.parquet as parquet

    table = parquet.read_table(path)
    rows = [list(row.values()) for row in table.to_pylist()]
    profile = _profile_rows(table.column_names, rows)
    for field, column_profile in zip(table.schema, profile.columns):
        column_profile.data_type = str(field.type)
    return profile


def profile_pdf(path: Path) -> DatasetProfile:
    from pypdf import PdfReader

    reader = PdfReader(path)
    page_text = [page.extract_text() or "" for page in reader.pages]
    issues: list[DataQualityIssue] = []
    warnings: list[str] = []
    if not any(text.strip() for text in page_text):
        warnings.append("No digital text was found; this PDF may require OCR.")
        issues.append(DataQualityIssue(code="ocr_required", severity="warning", message="The document contains no extractable text and likely requires OCR."))
    return DatasetProfile(dataset_kind="financial_document", page_count=len(reader.pages), extracted_character_count=sum(len(text) for text in page_text), warnings=warnings, quality_issues=issues)


def profile_file(path: Path, file_format: str) -> DatasetProfile:
    profilers = {"csv": profile_csv, "xlsx": profile_xlsx, "parquet": profile_parquet, "pdf": profile_pdf}
    return profilers[file_format](path)
